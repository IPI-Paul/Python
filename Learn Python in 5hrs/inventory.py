import openpyxl
import pandas
import os

inv_file = openpyxl.load_workbook("xlsx/inventory.xlsx")
product_list = inv_file["Sheet1"]


def exercise1():
    products_per_supplier = {}

    for product_row in range(2, product_list.max_row + 1):
        supplier_name = product_list.cell(product_row, 4).value

        # calculation for number of products per supplier
        """
        if supplier_name in products_per_supplier:
            current_num_products = products_per_supplier[supplier_name]
            products_per_supplier[supplier_name] = current_num_products + 1
        else
            print('adding a new supplier')
            products_per_supplier[supplier_name] = 1
        """
        products_per_supplier[supplier_name] = products_per_supplier.get(supplier_name, 0) + 1
    print(products_per_supplier)


def exercise2():
    products_per_supplier = {}
    total_value_per_supplier = {}

    for product_row in range(2, product_list.max_row + 1):
        supplier_name = product_list.cell(product_row, 4).value
        inventory = product_list.cell(product_row, 2).value
        price = product_list.cell(product_row, 3).value

        # calculation for number of products per supplier
        products_per_supplier[supplier_name] = products_per_supplier.get(supplier_name, 0) + 1

        # calculate total value of inventory per supplier
        """
        if supplier_name in total_value_per_supplier:
            current_total_value = total_value_per_supplier.get(supplier_name)
            total_value_per_supplier[supplier_name] = current_total_value + inventory * price
        else:
            total_value_per_supplier[supplier_name] = inventory * price
        """
        total_value_per_supplier[supplier_name] = total_value_per_supplier.get(supplier_name, 0) + inventory * price
    print(products_per_supplier)
    print(dict(map(dict.popitem, map(lambda x: {x: format(float(total_value_per_supplier[x]), ',.2f')},
                                     total_value_per_supplier))))


def exercise3():
    products_per_supplier = {}
    total_value_per_supplier = {}
    products_under_10_inv = {}

    for product_row in range(2, product_list.max_row + 1):
        supplier_name = product_list.cell(product_row, 4).value
        inventory = product_list.cell(product_row, 2).value
        price = product_list.cell(product_row, 3).value
        product_num = product_list.cell(product_row, 1).value

        # calculation for number of products per supplier
        products_per_supplier[supplier_name] = products_per_supplier.get(supplier_name, 0) + 1

        # calculate total value of inventory per supplier
        total_value_per_supplier[supplier_name] = total_value_per_supplier.get(supplier_name, 0) + inventory * price

        # logic products with inventory less than 10
        if inventory < 10:
            products_under_10_inv[int(product_num)] = int(inventory)

    print(products_under_10_inv)


def exercise4():
    products_per_supplier = {}
    total_value_per_supplier = {}
    products_under_10_inv = {}
    product_list.cell(1, 5).value = "Total Inventory Price"

    for product_row in range(2, product_list.max_row + 1):
        supplier_name = product_list.cell(product_row, 4).value
        inventory = product_list.cell(product_row, 2).value
        price = product_list.cell(product_row, 3).value
        product_num = product_list.cell(product_row, 1).value
        inventory_price = product_list.cell(product_row, 5)

        # calculation for number of products per supplier
        products_per_supplier[supplier_name] = products_per_supplier.get(supplier_name, 0) + 1

        # calculate total value of inventory per supplier
        total_value_per_supplier[supplier_name] = total_value_per_supplier.get(supplier_name, 0) + inventory * price

        # logic products with inventory less than 10
        if inventory < 10:
            products_under_10_inv[int(product_num)] = int(inventory)

        # add value for total inventory price
        inventory_price.value = format(inventory * price, ".2f")

    print(products_per_supplier)
    print(dict(map(dict.popitem, map(lambda x: {x: format(float(total_value_per_supplier[x]), ',.2f')},
                                     total_value_per_supplier))))
    print(products_under_10_inv)

    inv_file.save("xlsx/inventory_with_total_value.xlsx")
    df = pandas.read_excel("xlsx/inventory_with_total_value.xlsx")
    df.to_csv("csv/inventory_with_total_value.csv", index=None, header=True)
    os.remove("xlsx/inventory_with_total_value.xlsx")


user_input = None
while user_input not in ["", "exit"]:
    msg = "> " if user_input is not None else """
Please enter the exercise number to run!
1) Number of Products per Supplier
2) Total Value of Products per Supplier
3) Product IDs with an Inventory Under 10
4) Add a 'Total Inventory Price' Column and Save to csv File
h) Show Help List
type exit or just enter to quit
> """
    user_input = input(msg)
    if user_input in ["", "exit"]:
        print("All done, bye ;D")
        continue
    elif user_input == "h":
        user_input = None
        continue
    elif user_input not in ["1", "2", "3", "4"]:
        print(f"Your entry {user_input} is not one of the specified values!{msg}")
        user_input = None
        continue
    eval(f"exercise{user_input}()")
